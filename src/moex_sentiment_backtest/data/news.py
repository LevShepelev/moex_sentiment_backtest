from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable, Optional, Sequence

from moex_sentiment_backtest.logging_config import get_logger

log = get_logger(__name__)

_TICKER_RE = re.compile(r"^[A-Z0-9]{3,6}$")


def _norm_sent(v) -> int:
    """Normalize sentiment cell to int in [-2,2]. Unknown -> 0."""
    if v is None:
        return 0
    if isinstance(v, bool):
        return 0
    if isinstance(v, int):
        s = v
    elif isinstance(v, float):
        s = int(v)
    elif isinstance(v, str):
        t = v.strip()
        if t == "+":
            return 1
        if t == "-":
            return -1
        if t == "":
            return 0
        try:
            s = int(float(t))
        except Exception:
            return 0
    else:
        try:
            s = int(v)
        except Exception:
            return 0

    if s > 2:
        s = 2
    elif s < -2:
        s = -2
    return int(s)


def _norm_ts(v) -> Optional[datetime]:
    """
    Normalize timestamp value from Excel cell to python datetime (naive).
    Accepts: datetime, date, ISO-like str. Unknown -> None.
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, time.min)
    if isinstance(v, str):
        t = v.strip()
        if not t:
            return None
        # try ISO / "YYYY-mm-dd HH:MM:SS" / "YYYY-mm-dd"
        try:
            return datetime.fromisoformat(t)
        except Exception:
            pass
        # very common excel string format
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(t, fmt)
            except Exception:
                continue
        return None
    # fallback
    try:
        return datetime.fromisoformat(str(v))
    except Exception:
        return None


def _infer_date_column(cols: list[str]) -> str:
    norm_map = {str(c).strip().lower(): i for i, c in enumerate(cols)}
    for cand in (
        "date",
        "datetime",
        "timestamp",
        "time",
        "ts",
        "posted_at",
        "created_at",
        "published_at",
        "дата",
        "время",
        "дата_время",
        "дата-время",
    ):
        if cand in norm_map:
            return cols[norm_map[cand]]
    raise ValueError(f"Could not infer date column. First columns: {cols[:50]}")


def _parse_sheet_arg(sheet_name: Optional[str | int], wb_sheetnames: Sequence[str]) -> list[str]:
    """
    Return list of sheet names to process.
    sheet_name can be:
      - None -> []
      - int -> index into workbook sheets
      - str -> single name or comma-separated names
    """
    if sheet_name is None:
        return []
    if isinstance(sheet_name, int):
        idx = int(sheet_name)
        if idx < 0 or idx >= len(wb_sheetnames):
            raise ValueError(f"sheet_name index {idx} out of range. Available: {list(wb_sheetnames)}")
        return [wb_sheetnames[idx]]
    s = str(sheet_name).strip()
    if not s:
        return []
    # allow "Sheet1,Sheet2"
    parts = [p.strip() for p in s.split(",") if p.strip()]
    return parts


def prepare_events(
    news_path: Path,
    out_path: Path,
    keep_text: bool = False,
    ticker_columns: Optional[list[str]] = None,
    date_column: Optional[str] = None,
    sheet_name: Optional[str | int] = None,
    all_sheets: bool = False,
    flush_every_events: int = 500_000,
) -> None:
    """
    Stream wide .xlsx and write long events parquet without loading whole file.

    - If all_sheets=True: reads ALL sheets and concatenates events.
    - If sheet_name is "Sheet1,Sheet2": reads listed sheets.
    - If sheet_name is int: reads sheet by index.
    - Else reads active sheet only.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import load_workbook
    import pyarrow as pa
    import pyarrow.parquet as pq

    wb = load_workbook(news_path, read_only=True, data_only=True)

    # choose sheets
    wb_sheetnames = list(wb.sheetnames)
    chosen: list[str]
    if all_sheets:
        chosen = wb_sheetnames
    else:
        chosen = _parse_sheet_arg(sheet_name, wb_sheetnames)
        if not chosen:
            chosen = [wb.active.title]

    log.info("Reading sheets: %s", chosen)

    # Parquet schema is fixed across sheets
    meta_names: list[str] = []
    if keep_text:
        meta_names = ["id", "text", "post_link", "message_id", "channel_link", "Scenario", "LLM_Output"]

    schema_fields = [
        pa.field("ts", pa.timestamp("us")),
        pa.field("ticker", pa.string()),
        pa.field("sentiment", pa.int8()),
    ] + [pa.field(m, pa.string()) for m in meta_names]

    schema = pa.schema(schema_fields)
    writer = pq.ParquetWriter(out_path, schema, compression="zstd", use_dictionary=True)

    ts_buf: list[datetime] = []
    ticker_buf: list[str] = []
    sent_buf: list[int] = []
    meta_bufs = {m: [] for m in meta_names}

    written = 0

    def flush() -> None:
        nonlocal written
        if not ts_buf:
            return
        arrays = {
            "ts": pa.array(ts_buf, type=pa.timestamp("us")),
            "ticker": pa.array(ticker_buf, type=pa.string()),
            "sentiment": pa.array(sent_buf, type=pa.int8()),
        }
        for m in meta_names:
            arrays[m] = pa.array(meta_bufs[m], type=pa.string())

        table = pa.Table.from_pydict(arrays, schema=schema)
        writer.write_table(table, row_group_size=min(len(ts_buf), 250_000))
        written += len(ts_buf)

        ts_buf.clear()
        ticker_buf.clear()
        sent_buf.clear()
        for m in meta_names:
            meta_bufs[m].clear()

    # process each sheet
    for sh in chosen:
        ws = wb[sh]

        # header
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        cols = [("" if c is None else str(c)) for c in header_row]

        # date col
        effective_date_col = date_column or _infer_date_column(cols)
        # locate date index (normalized match)
        norm_map = {str(c).strip().lower(): i for i, c in enumerate(cols)}
        dnorm = str(effective_date_col).strip().lower()
        if dnorm not in norm_map:
            raise ValueError(f"date_column='{effective_date_col}' not found in sheet '{sh}'. First cols: {cols[:50]}")
        date_idx = norm_map[dnorm]

        # ticker cols for this sheet
        if ticker_columns is None:
            tickers_this = [c for c in cols if c and c != effective_date_col and _TICKER_RE.fullmatch(c.strip() or "")]
        else:
            # use provided tickers, but only those present in this sheet
            tickers_this = [t for t in ticker_columns if t in cols]

        if not tickers_this:
            raise ValueError(f"Could not infer ticker columns in sheet '{sh}'. Pass --ticker_columns explicitly.")

        ticker_idx = [(tkr, cols.index(tkr)) for tkr in tickers_this]

        # meta indices (only if keep_text)
        meta_idx: list[tuple[str, int]] = []
        if meta_names:
            col_index = {c: i for i, c in enumerate(cols)}
            for m in meta_names:
                if m in col_index:
                    meta_idx.append((m, col_index[m]))

        log.info(
            "Sheet '%s': rows streaming, tickers=%d, date_col='%s'",
            sh,
            len(ticker_idx),
            effective_date_col,
        )

        # iterate rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            # ts
            ts_raw = row[date_idx] if date_idx < len(row) else None
            ts = _norm_ts(ts_raw)
            if ts is None:
                continue

            # meta once per row
            meta_vals = {}
            if meta_idx:
                for m, mi in meta_idx:
                    v = row[mi] if mi < len(row) else None
                    meta_vals[m] = "" if v is None else str(v)

            # emit only non-zero sentiments
            for tkr, ti in ticker_idx:
                v = row[ti] if ti < len(row) else None
                s = _norm_sent(v)
                if s == 0:
                    continue
                ts_buf.append(ts)
                ticker_buf.append(tkr)
                sent_buf.append(s)
                for m in meta_names:
                    meta_bufs[m].append(meta_vals.get(m, ""))

            if len(ts_buf) >= flush_every_events:
                flush()

    flush()
    writer.close()
    wb.close()
    log.info("Saved %d events -> %s", written, str(out_path))